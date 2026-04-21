"""
parse_html_reports.py
─────────────────────────────────────────────────────────────────────────────
hybrid_report_*.html 파일 전체를 파싱해 통계 데이터를 추출하고
Excel(.xlsx)로 내보낸다.

사용:
    python3 parse_html_reports.py                     # reports/ 전체 파싱
    python3 parse_html_reports.py --from 2026-04-01   # 날짜 필터
    python3 parse_html_reports.py --from 2026-04-01 --to 2026-04-08
    python3 parse_html_reports.py --out /tmp/stats.xlsx
    python3 parse_html_reports.py --print             # 콘솔 요약만
"""
from __future__ import annotations

import argparse
import glob
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from html.parser import HTMLParser
from pathlib import Path


# ─── HTML 파서 ────────────────────────────────────────────────────────────────

class ReportParser(HTMLParser):
    """hybrid_report_*.html 에서 구조화된 데이터를 추출한다."""

    def __init__(self):
        super().__init__()
        self._tag_stack: list[str] = []
        self._cls_stack: list[str] = []
        self._skip_depth = 0          # script/style/audio/source/img 안에 있으면 ++

        # 추출 결과
        self.created_at: str | None = None          # "2026-04-01 14:23:45"
        self.scenario: str | None = None
        self.android_app: str | None = None
        self.ios_app: str | None = None
        self.android_device: str | None = None
        self.ios_device: str | None = None
        self.android_os: str | None = None
        self.ios_os: str | None = None

        # 총합 음단절 (tfoot 의 badge-num 두 개)
        self.total_android_dropouts: int | None = None
        self.total_ios_dropouts: int | None = None

        # MOS 행별 데이터: list of dict
        self.mos_rows: list[dict] = []

        # ── 내부 상태 ──
        self._env_row_mode = False       # <div class="env-block"> 안
        self._env_label: str | None = None
        self._in_env_val = False

        self._in_tfoot = False
        self._tfoot_badge_count = 0      # tfoot 안 badge-num 수집 순서

        self._in_mos_table = False       # <table class="mos-table"> 안
        self._mos_row: dict | None = None
        self._mos_col_idx = -1           # 현재 td 인덱스 (0=label,1=ios신호,2=and신호,3=iOS MOS,4=AND MOS)
        self._mos_pending_os: str | None = None  # 마지막 mos-os-badge가 "ios" or "and"

        self._cur_cls = ""              # 현재 태그의 class (handle_data 에서 참조)

    # ── 태그 진입/종료 ──
    def handle_starttag(self, tag: str, attrs: list):
        attrs_d = dict(attrs)
        cls = attrs_d.get("class", "")
        self._tag_stack.append(tag)
        self._cls_stack.append(cls)
        self._cur_cls = cls

        # img, source 는 void/self-closing → endtag 없음 → skip_depth 제외
        if tag in ("script", "style", "audio"):
            self._skip_depth += 1
            return

        # env-block 시작
        if tag == "div" and "env-block" in cls:
            self._env_row_mode = True

        # env-table td (첫 번째 td = 라벨)
        if self._env_row_mode and tag == "td":
            self._env_label = None
            self._in_env_val = False

        # tfoot 감지
        if tag == "tfoot":
            self._in_tfoot = True
            self._tfoot_badge_count = 0

        # mos-table 감지
        if tag == "table" and "mos-table" in cls:
            self._in_mos_table = True

        # mos-table 내부 tr
        if self._in_mos_table and tag == "tr":
            self._mos_row = {}
            self._mos_col_idx = -1

        # mos-table 내부 td
        if self._in_mos_table and tag == "td":
            self._mos_col_idx += 1

    def handle_endtag(self, tag: str):
        if tag in ("script", "style", "audio"):
            self._skip_depth = max(0, self._skip_depth - 1)

        if self._tag_stack and self._tag_stack[-1] == tag:
            self._tag_stack.pop()
            self._cls_stack.pop()
        self._cur_cls = self._cls_stack[-1] if self._cls_stack else ""

        if tag == "tfoot":
            self._in_tfoot = False

        if tag == "table":
            self._in_mos_table = False

        # mos row 완료
        if self._in_mos_table and tag == "tr" and self._mos_row is not None:
            if self._mos_row:                # 빈 헤더 행 제외
                self.mos_rows.append(self._mos_row)
            self._mos_row = None
            self._mos_col_idx = -1

        if tag == "div" and self._env_row_mode:
            # env-block 종료 판단: 가장 바깥 div만 노출하면 충분
            pass

    def handle_data(self, data: str):
        if self._skip_depth > 0:
            return
        if "base64" in data:
            return
        data_s = data.strip()
        if not data_s:
            return

        cls = self._cur_cls

        # ── 생성 일시 ──────────────────────────────────────────────────────
        if self.created_at is None and data_s.startswith("생성:"):
            m = re.search(r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})", data_s)
            if m:
                self.created_at = m.group(1)
            m2 = re.search(r"📁\s*(.+?)(?:\s*$|\s*&)", data_s)
            if m2:
                self.scenario = m2.group(1).strip()

        # ── env-table 파싱 ─────────────────────────────────────────────────
        if self._env_row_mode:
            # env-table 의 첫 번째 td 는 라벨, 두 번째 td (b 태그) 는 값
            tag = self._tag_stack[-1] if self._tag_stack else ""
            if tag == "td" and self._env_label is None:
                self._env_label = data_s
            elif tag in ("b", "td") and self._env_label:
                lbl = self._env_label.rstrip(":")
                if "시나리오" in lbl:
                    self.scenario = self.scenario or data_s
                elif "Android 앱" in lbl:
                    self.android_app = data_s
                elif "iOS 앱" in lbl:
                    self.ios_app = data_s
                elif "Android 단말" in lbl:
                    self.android_device = data_s
                elif "iOS 단말" in lbl:
                    self.ios_device = data_s
                elif "Android OS" in lbl:
                    self.android_os = data_s
                elif "iOS OS" in lbl:
                    self.ios_os = data_s
                self._env_label = None

        # ── tfoot badge-num (총 음단절) ────────────────────────────────────
        if self._in_tfoot and "badge-num" in cls:
            try:
                val = int(data_s)
                if self._tfoot_badge_count == 0:
                    self.total_android_dropouts = val
                elif self._tfoot_badge_count == 1:
                    self.total_ios_dropouts = val
                self._tfoot_badge_count += 1
            except ValueError:
                pass

        # ── MOS 테이블 ────────────────────────────────────────────────────
        if self._in_mos_table and self._mos_row is not None:
            col = self._mos_col_idx

            # 라벨 열 (col=0)
            if col == 0 and "mos-label" in cls:
                self._mos_row["label"] = data_s

            # iOS 신호 열 (col=1) — SNR / RMS
            if col == 1:
                m = re.search(r"SNR\s+([\d.]+)\s*dB", data_s)
                if m:
                    self._mos_row["ios_snr_db"] = float(m.group(1))
                m2 = re.search(r"RMS\s+([\d.]+)", data_s)
                if m2:
                    self._mos_row["ios_rms"] = float(m2.group(1))

            # Android 신호 열 (col=2)
            if col == 2:
                m = re.search(r"SNR\s+([\d.]+)\s*dB", data_s)
                if m:
                    self._mos_row["android_snr_db"] = float(m.group(1))
                m2 = re.search(r"RMS\s+([\d.]+)", data_s)
                if m2:
                    self._mos_row["android_rms"] = float(m2.group(1))

            # iOS MOS 열 (col=3)
            if col == 3:
                if "mos-score" in cls:
                    try:
                        self._mos_row["ios_mos"] = float(data_s)
                    except ValueError:
                        pass
                if "mos-grade" in cls:
                    self._mos_row["ios_mos_grade"] = data_s

            # Android MOS 열 (col=4)
            if col == 4:
                if "mos-score" in cls:
                    try:
                        self._mos_row["android_mos"] = float(data_s)
                    except ValueError:
                        pass
                if "mos-grade" in cls:
                    self._mos_row["android_mos_grade"] = data_s


def parse_report(path: str) -> dict | None:
    """단일 HTML 파일에서 데이터를 추출해 dict로 반환. 실패 시 None."""
    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            content = f.read()
    except OSError:
        return None

    p = ReportParser()
    try:
        p.feed(content)
    except Exception:
        return None

    # 파일명에서 unix timestamp → 날짜 fallback
    fname = os.path.basename(path)
    m_ts = re.search(r"hybrid_report_(\d+)\.html", fname)
    file_dt: str | None = None
    if m_ts:
        try:
            file_dt = datetime.fromtimestamp(int(m_ts.group(1))).strftime("%Y-%m-%d %H:%M:%S")
        except (ValueError, OSError):
            pass

    created = p.created_at or file_dt
    if not created:
        return None

    # MOS 집계 (여러 음원의 평균)
    ios_mos_vals = [r["ios_mos"] for r in p.mos_rows if "ios_mos" in r]
    and_mos_vals = [r["android_mos"] for r in p.mos_rows if "android_mos" in r]
    ios_snr_vals = [r["ios_snr_db"] for r in p.mos_rows if "ios_snr_db" in r]
    and_snr_vals = [r["android_snr_db"] for r in p.mos_rows if "android_snr_db" in r]

    def avg(lst):
        return round(sum(lst) / len(lst), 4) if lst else None

    return {
        "file": fname,
        "path": path,
        "created_at": created,
        "date": created[:10],
        "scenario": p.scenario,
        "android_app": p.android_app,
        "ios_app": p.ios_app,
        "android_device": p.android_device,
        "ios_device": p.ios_device,
        "android_os": p.android_os,
        "ios_os": p.ios_os,
        "android_dropouts": p.total_android_dropouts,
        "ios_dropouts": p.total_ios_dropouts,
        "ios_mos": avg(ios_mos_vals),
        "android_mos": avg(and_mos_vals),
        "ios_snr_db": avg(ios_snr_vals),
        "android_snr_db": avg(and_snr_vals),
        "mos_row_count": len(p.mos_rows),
        "mos_rows": p.mos_rows,     # 음원별 상세
    }


# ─── 집계 유틸 ────────────────────────────────────────────────────────────────

def _week_label(date_str: str) -> str:
    """'2026-04-01' → '2026-W14'"""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    return dt.strftime("%Y-W%W")

def _month_label(date_str: str) -> str:
    return date_str[:7]   # '2026-04'

def _year_label(date_str: str) -> str:
    return date_str[:4]   # '2026'

def aggregate(records: list[dict], key_fn) -> list[dict]:
    """records를 key_fn으로 그룹화해 집계 행 목록 반환."""
    groups: dict[str, list[dict]] = defaultdict(list)
    for r in records:
        groups[key_fn(r)].append(r)

    rows = []
    for period in sorted(groups):
        grp = groups[period]
        n = len(grp)

        def agg_avg(field):
            vals = [r[field] for r in grp if r.get(field) is not None]
            return round(sum(vals) / len(vals), 4) if vals else None

        def agg_sum(field):
            vals = [r[field] for r in grp if r.get(field) is not None]
            return sum(vals) if vals else None

        def agg_pct(drop_field):
            """음단절 발생 리포트 비율 (dropouts > 0 인 파일 수 / 전체)"""
            vals = [r[drop_field] for r in grp if r.get(drop_field) is not None]
            if not vals:
                return None
            nonzero = sum(1 for v in vals if v > 0)
            return round(nonzero / len(vals) * 100, 1)

        row = {
            "기간": period,
            "리포트 수": n,
            "iOS MOS 평균": agg_avg("ios_mos"),
            "Android MOS 평균": agg_avg("android_mos"),
            "iOS SNR 평균(dB)": agg_avg("ios_snr_db"),
            "Android SNR 평균(dB)": agg_avg("android_snr_db"),
            "iOS 음단절 합계": agg_sum("ios_dropouts"),
            "Android 음단절 합계": agg_sum("android_dropouts"),
            "iOS 음단절 발생율(%)": agg_pct("ios_dropouts"),
            "Android 음단절 발생율(%)": agg_pct("android_dropouts"),
        }
        rows.append(row)
    return rows


# ─── Excel 내보내기 ───────────────────────────────────────────────────────────

def to_excel(records: list[dict], out_path: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[!] openpyxl 없음 → pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    HEADER_FILL  = PatternFill("solid", fgColor="1E2D4A")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    GOOD_FILL    = PatternFill("solid", fgColor="1A2A1A")
    WARN_FILL    = PatternFill("solid", fgColor="2A1A00")
    BAD_FILL     = PatternFill("solid", fgColor="2A0000")

    def _mos_fill(v):
        if v is None: return None
        if v >= 4.0: return GOOD_FILL
        if v >= 3.0: return None
        if v >= 2.5: return WARN_FILL
        return BAD_FILL

    def write_sheet(ws, headers: list[str], rows: list[dict], col_widths: list[int]):
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(rows, 2):
            for ci, h in enumerate(headers, 1):
                val = row.get(h)
                cell = ws.cell(ri, ci, val)
                cell.alignment = Alignment(horizontal="center" if isinstance(val, (int, float)) else "left")
                # MOS 컬럼 색상
                if "MOS" in h and isinstance(val, float):
                    fill = _mos_fill(val)
                    if fill:
                        cell.fill = fill
                # 음단절 발생 강조
                if "음단절" in h and isinstance(val, (int, float)) and val > 0:
                    cell.fill = WARN_FILL
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"

    # ── 시트1: 전체 리포트 목록 ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "전체 리포트"
    h1 = ["날짜", "생성시각", "시나리오", "Android 앱", "iOS 앱",
          "Android 단말", "iOS 단말", "Android OS", "iOS OS",
          "iOS MOS", "Android MOS", "iOS SNR(dB)", "Android SNR(dB)",
          "iOS 음단절 수", "Android 음단절 수", "음원 수", "파일명"]
    rows1 = [{
        "날짜": r["date"],
        "생성시각": r["created_at"],
        "시나리오": r.get("scenario") or "",
        "Android 앱": r.get("android_app") or "",
        "iOS 앱": r.get("ios_app") or "",
        "Android 단말": r.get("android_device") or "",
        "iOS 단말": r.get("ios_device") or "",
        "Android OS": r.get("android_os") or "",
        "iOS OS": r.get("ios_os") or "",
        "iOS MOS": r.get("ios_mos"),
        "Android MOS": r.get("android_mos"),
        "iOS SNR(dB)": r.get("ios_snr_db"),
        "Android SNR(dB)": r.get("android_snr_db"),
        "iOS 음단절 수": r.get("ios_dropouts"),
        "Android 음단절 수": r.get("android_dropouts"),
        "음원 수": r.get("mos_row_count", 0),
        "파일명": r["file"],
    } for r in records]
    write_sheet(ws1, h1, rows1, [12, 20, 20, 18, 18, 24, 24, 14, 14, 10, 10, 12, 12, 12, 14, 6, 40])

    # ── 시트2~4: 기간별 집계 ───────────────────────────────────────────────
    period_defs = [
        ("일별 통계",  lambda r: r["date"]),
        ("주별 통계",  lambda r: _week_label(r["date"])),
        ("월별 통계",  lambda r: _month_label(r["date"])),
        ("연도별 통계", lambda r: _year_label(r["date"])),
    ]
    agg_headers = [
        "기간", "리포트 수",
        "iOS MOS 평균", "Android MOS 평균",
        "iOS SNR 평균(dB)", "Android SNR 평균(dB)",
        "iOS 음단절 합계", "Android 음단절 합계",
        "iOS 음단절 발생율(%)", "Android 음단절 발생율(%)",
    ]
    agg_widths = [14, 10, 14, 16, 16, 18, 14, 16, 18, 20]

    for sheet_name, key_fn in period_defs:
        ws = wb.create_sheet(sheet_name)
        agg_rows = aggregate(records, key_fn)
        write_sheet(ws, agg_headers, agg_rows, agg_widths)

    # ── 시트5: 음원별 MOS 상세 ──────────────────────────────────────────────
    ws5 = wb.create_sheet("음원별 MOS 상세")
    h5 = ["날짜", "생성시각", "음원 라벨",
          "iOS MOS", "iOS MOS 등급", "iOS SNR(dB)", "iOS RMS",
          "Android MOS", "Android MOS 등급", "Android SNR(dB)", "Android RMS",
          "파일명"]
    rows5 = []
    for r in records:
        for mr in r.get("mos_rows", []):
            rows5.append({
                "날짜": r["date"],
                "생성시각": r["created_at"],
                "음원 라벨": mr.get("label", ""),
                "iOS MOS": mr.get("ios_mos"),
                "iOS MOS 등급": mr.get("ios_mos_grade", ""),
                "iOS SNR(dB)": mr.get("ios_snr_db"),
                "iOS RMS": mr.get("ios_rms"),
                "Android MOS": mr.get("android_mos"),
                "Android MOS 등급": mr.get("android_mos_grade", ""),
                "Android SNR(dB)": mr.get("android_snr_db"),
                "Android RMS": mr.get("android_rms"),
                "파일명": r["file"],
            })
    write_sheet(ws5, h5, rows5, [12, 20, 20, 10, 12, 12, 10, 12, 14, 14, 10, 40])

    wb.save(out_path)
    print(f"✅ Excel 저장: {out_path}")


# ─── 콘솔 요약 출력 ───────────────────────────────────────────────────────────

def print_summary(records: list[dict]):
    print(f"\n{'='*64}")
    print(f"  파싱된 리포트: {len(records)}개")
    if not records:
        return

    dates = sorted(r["date"] for r in records)
    print(f"  기간: {dates[0]} ~ {dates[-1]}")
    print()

    # 전체 집계
    ios_mos  = [r["ios_mos"] for r in records if r.get("ios_mos") is not None]
    and_mos  = [r["android_mos"] for r in records if r.get("android_mos") is not None]
    ios_snr  = [r["ios_snr_db"] for r in records if r.get("ios_snr_db") is not None]
    and_snr  = [r["android_snr_db"] for r in records if r.get("android_snr_db") is not None]
    ios_drop = [r["ios_dropouts"] for r in records if r.get("ios_dropouts") is not None]
    and_drop = [r["android_dropouts"] for r in records if r.get("android_dropouts") is not None]

    def avg(lst): return sum(lst)/len(lst) if lst else None
    def fmt(v, d=3): return f"{v:.{d}f}" if v is not None else "—"

    print(f"  iOS MOS      평균: {fmt(avg(ios_mos))}  |  Android MOS      평균: {fmt(avg(and_mos))}")
    print(f"  iOS SNR      평균: {fmt(avg(ios_snr),1)} dB  |  Android SNR      평균: {fmt(avg(and_snr),1)} dB")
    print(f"  iOS 음단절 총합: {sum(ios_drop)}건 ({sum(1 for v in ios_drop if v>0)}/{len(ios_drop)}파일 발생)")
    print(f"  AOS 음단절 총합: {sum(and_drop)}건 ({sum(1 for v in and_drop if v>0)}/{len(and_drop)}파일 발생)")

    # 일별 요약
    print(f"\n{'─'*64}")
    print(f"  {'날짜':12} {'리포트':>6} {'iOS MOS':>9} {'AOS MOS':>9} {'iOS 음단절':>10} {'AOS 음단절':>10}")
    print(f"{'─'*64}")
    for row in aggregate(records, lambda d: d["date"]):
        p = row["기간"]
        n = row["리포트 수"]
        im  = fmt(row["iOS MOS 평균"])
        am  = fmt(row["Android MOS 평균"])
        id_ = row["iOS 음단절 합계"] if row["iOS 음단절 합계"] is not None else "—"
        ad_ = row["Android 음단절 합계"] if row["Android 음단절 합계"] is not None else "—"
        ir  = row["iOS 음단절 발생율(%)"]
        ar  = row["Android 음단절 발생율(%)"]
        id_s = f"{id_}건({ir}%)" if ir is not None else f"{id_}건"
        ad_s = f"{ad_}건({ar}%)" if ar is not None else f"{ad_}건"
        print(f"  {p:12} {n:>6}   {im:>7}   {am:>7}   {id_s:>12}   {ad_s:>12}")
    print(f"{'='*64}\n")


# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="hybrid_report HTML 통계 파서")
    ap.add_argument("--dir", default="/Users/qabulls/Documents/sound/reports",
                    help="리포트 디렉토리 (default: reports/)")
    ap.add_argument("--from", dest="from_date", default=None,
                    help="시작 날짜 YYYY-MM-DD")
    ap.add_argument("--to", dest="to_date", default=None,
                    help="종료 날짜 YYYY-MM-DD")
    ap.add_argument("--out", default=None,
                    help="Excel 출력 경로 (기본: ~/Desktop/html_report_stats_YYYY-MM-DD.xlsx)")
    ap.add_argument("--print", dest="print_only", action="store_true",
                    help="콘솔 요약만 출력 (Excel 생성 안함)")
    args = ap.parse_args()

    # 파일 탐색
    pattern = os.path.join(args.dir, "**", "hybrid_report_*.html")
    all_files = sorted(glob.glob(pattern, recursive=True))
    print(f"[*] HTML 파일 {len(all_files)}개 발견")

    # 파싱
    records = []
    errors  = 0
    for fpath in all_files:
        rec = parse_report(fpath)
        if rec is None:
            errors += 1
            continue
        records.append(rec)

    # 날짜 필터
    if args.from_date:
        records = [r for r in records if r["date"] >= args.from_date]
    if args.to_date:
        records = [r for r in records if r["date"] <= args.to_date]

    # 날짜순 정렬
    records.sort(key=lambda r: r["created_at"])

    print(f"[*] 파싱 성공: {len(records)}개  |  실패: {errors}개")

    print_summary(records)

    if args.print_only:
        return

    out = args.out
    if out is None:
        today = datetime.now().strftime("%Y-%m-%d")
        out = os.path.expanduser(f"~/Desktop/html_report_stats_{today}.xlsx")

    to_excel(records, out)


if __name__ == "__main__":
    main()
